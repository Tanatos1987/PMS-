"""
ClockPMS+ RPA Automation + Excel Processing  (REVIZIRANA / РАБОТЕЩА ВЕРСИЯ)
==========================================================================
Поправена версия, написана срещу РЕАЛНИЯ интерфейс на ClockPMS+ (Quasar SPA).

Какво беше счупено в старата версия и е поправено тук:
  • ClockPMS+ е Quasar/Vue приложение и НЕ дава фиксирани id-та на полетата
    (id-тата са случайни GUID-и, напр. f_0acf0979-...). Затова всички
    `set_date_field(driver, "creation_from", ...)` и търсения по `By.ID(...)`
    хвърляха NoSuchElement и скриптът спираше веднага след логин.
    → Сега полетата се намират ПО ЕТИКЕТ (label text).
  • Навигацията през меню беше чуплива (XPath-и хващаха <body>).
    → Сега навигацията е ДИРЕКТНО ПО URL (намерено от реалния интерфейс).
  • „Стойност" опцията беше „Нетна стойност" — реално е „Нето".
  • Експортът е иконата за сваляне (⬇) → меню → „Excel".
  • Невалиден XPath в проверката за логин — поправен.
  • Добавена е диагностика: при всяка грешка се пази екранна снимка + HTML
    в папка debug/, за да се вижда точно къде/защо спира.

Инсталация:
    pip install selenium pandas openpyxl webdriver-manager

Преди стартиране:
    1. Попълни PASSWORD (и при нужда USERNAME) по-долу.
    2. Постави master Excel файла в WORK_DIR (виж MASTER_FILE).
    3. Стартирай: python clockpms_automation.py
"""

import os
import re
import time
import shutil
import traceback
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# ─────────────────────────────────────────────────────────────
#  КОНФИГУРАЦИЯ — попълни тук
# ─────────────────────────────────────────────────────────────
CLOCKPMS_BASE = ""
HOTEL_ID      = "14539"          # от URL-а след логин: /hotel/14539/...

USERNAME = ""        # използва се само ако трябва смяна на потребител
PASSWORD = ""           # <-- ПОПЪЛНИ реалната парола
SERIAL_KEY = ""  # нужен само при нов профил

# Папки
WORK_DIR     = Path(__file__).parent
DOWNLOAD_DIR = WORK_DIR / "downloads"
DEBUG_DIR    = WORK_DIR / "debug"
DOWNLOAD_DIR.mkdir(exist_ok=True)
DEBUG_DIR.mkdir(exist_ok=True)

# ── РЕЖИМ НА РАБОТА ─────────────────────────────────────────
# ATTACH_TO_RUNNING = True  → НАЙ-ЛЕСНИЯТ начин:
#   1) Стартираш "Start_Chrome.bat" (отваря Chrome + ClockPMS).
#   2) Логваш се РЪЧНО.
#   3) Пускаш този скрипт — той се „закача" за вече отворения Chrome
#      и поема нататък (търсене, анулации, справки, Excel).
#   Така няма нова станция, няма заключен профил, няма проблем с логина.
#
# ATTACH_TO_RUNNING = False → скриптът сам пуска Chrome и се логва
#   (по-чупливо; ползвай само ако горният режим не върви).
ATTACH_TO_RUNNING = True
DEBUG_PORT        = 9222

# Chrome профил (ползва се само при ATTACH_TO_RUNNING = False).
CHROME_USER_DATA = r"C:\Users\k.ivanova.TSMEGA\AppData\Local\Google\Chrome\User Data"
CHROME_PROFILE   = "Default"

# Master Excel файл
MASTER_FILE = WORK_DIR / "Обобщение.xlsx"        # при нужда смени на "Обобщена.xlsx"
OUTPUT_FILE = WORK_DIR / "Обобщение_Updated.xlsx"
SHEET_RES   = "Резервации 2026"

# Периоди (формат ДД.ММ.ГГГГ — конвертира се автоматично според полето)
NEW_RES_FROM = "13.06.2026"
NEW_RES_TO   = "19.06.2026"
CANCEL_FROM  = "01.10.2025"
CANCEL_TO    = "19.06.2026"
COMPANY_FROM = "01.05.2026"
COMPANY_TO   = "30.09.2026"

# Статуси за нови резервации (точните етикети от интерфейса)
NEW_RES_STATUSES = ["Очаквана", "Настанена", "Изпратена", "Непристигнала"]
CANCEL_STATUS    = "Анулирана"

# Дата-поле, по което се филтрира (етикет в „Разширено търсене")
DATE_FIELD_LABEL = "Създадена"     # алтернативи: "Анулирана", "Пристигане", ...

# Стойност в справката: реалните опции са само „Бруто" / „Нето"
REPORT_VALUE   = "Нето"
REPORT_GROUPBY = "Пристигане"

# Речник компания → лист в master файла
COMPANY_SHEET_MAP = {
    "":                                       "Sunny",
    "":                                  "Magelan",
    "":                                      "Novoton",
    "НОВОТОН ГАРАНЦИЯ - ЕООД":                             "Novoton",
    "ДЕСТИНЕЙШЪН ТУРИСТИК СЪРВИСИС - ЕООД":                "DTS",
    "ТУРИСТИЧЕСКА АГЕНЦИЯ СОЛВЕКС - ЕООД":                 "Solvex",
    "СЛР-ХОЛИДЕЙ-СЪРВИС ООД":                              "SLR",
    "NOWA ITAKA SPOLKA Z OGRANICZONA ODPOWIEDZIALNOSCIA":  "Itaka",
    "BOOKING":                                             "Booking",
    "АВАТАР ТУР - ООД":                                    "Avatar",
}

EUR_TO_BGN = 1.95583

# Кратки/дълги паузи
SHORT = 0.4
MED   = 1.2


# ═════════════════════════════════════════════════════════════
#  ИНФРАСТРУКТУРА
# ═════════════════════════════════════════════════════════════

def url(path: str) -> str:
    """Изгражда пълен SPA URL за даден път (без водеща наклонена черта)."""
    return f"{CLOCKPMS_BASE}/hotel/{HOTEL_ID}/{path.lstrip('/')}"


def attach_driver() -> webdriver.Chrome:
    """
    Закача се за ВЕЧЕ ОТВОРЕН Chrome, пуснат с --remote-debugging-port=DEBUG_PORT
    (виж Start_Chrome.bat). Не пуска нов Chrome и не пипа логина.
    """
    print(f"[CHROME] Закачам се за отворения Chrome (порт {DEBUG_PORT})...")
    opts = Options()
    opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{DEBUG_PORT}")
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    except Exception as e:
        raise RuntimeError(
            "Не успях да се закача за Chrome. Първо стартирай 'Start_Chrome.bat', "
            "влез в ClockPMS+, и чак тогава пусни скрипта.\n"
            f"(детайл: {e})")
    # пренасочва свалянията към нашата папка (prefs не важат при закачане)
    try:
        driver.execute_cdp_cmd("Page.setDownloadBehavior",
            {"behavior": "allow", "downloadPath": str(DOWNLOAD_DIR.resolve())})
    except Exception:
        pass
    return driver


def build_driver() -> webdriver.Chrome:
    """Създава Chrome driver с твоя реален профил (затваря текущия Chrome)."""
    import subprocess
    print("[CHROME] Затварям всички Chrome прозорци и почиствам lock файлове...")
    subprocess.run(["taskkill", "/F", "/T", "/IM", "chrome.exe"],       capture_output=True)
    subprocess.run(["taskkill", "/F", "/T", "/IM", "chromedriver.exe"], capture_output=True)
    time.sleep(3)
    # изчаква chrome.exe да приключи напълно
    for _ in range(10):
        r = subprocess.run(["tasklist", "/FI", "IMAGENAME eq chrome.exe"],
                           capture_output=True, text=True)
        if "chrome.exe" not in (r.stdout or ""):
            break
        time.sleep(1)

    profile = Path(CHROME_USER_DATA)
    for lock in [profile / "SingletonLock", profile / "SingletonSocket",
                 profile / "SingletonCookie", profile / "lockfile",
                 profile / CHROME_PROFILE / "LOCK", profile / CHROME_PROFILE / "lockfile"]:
        try:
            lock.unlink(missing_ok=True)
        except Exception:
            pass
    time.sleep(1)

    prefs = {
        "download.default_directory": str(DOWNLOAD_DIR.resolve()),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    opts = Options()
    opts.add_experimental_option("prefs", prefs)
    opts.add_argument(f"--user-data-dir={CHROME_USER_DATA}")
    opts.add_argument(f"--profile-directory={CHROME_PROFILE}")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--no-first-run")
    opts.add_argument("--no-default-browser-check")
    opts.add_argument("--disable-default-apps")
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument("--start-maximized")
    # Стабилност (предотвратява 'DevToolsActivePort file doesn't exist')
    opts.add_argument("--remote-debugging-port=9222")
    opts.add_argument("--disable-dev-shm-usage")
    # opts.add_argument("--headless=new")  # ← откоментирай за headless

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
    })
    return driver


def debug_dump(driver, name: str):
    """Пази екранна снимка + HTML на текущата страница (за диагностика)."""
    ts = datetime.now().strftime("%H%M%S")
    safe = re.sub(r"[^\w\-]+", "_", name)[:50]
    try:
        driver.save_screenshot(str(DEBUG_DIR / f"{ts}_{safe}.png"))
    except Exception:
        pass
    try:
        (DEBUG_DIR / f"{ts}_{safe}.html").write_text(
            driver.page_source, encoding="utf-8")
    except Exception:
        pass
    print(f"  [DEBUG] Запазих снимка+HTML: debug/{ts}_{safe}.*")


# ═════════════════════════════════════════════════════════════
#  ОБЩИ UI ПОМОЩНИЦИ (Quasar)
# ═════════════════════════════════════════════════════════════

def _copy_to_clipboard(text: str):
    """Копира текст в Windows клипборда (работи с $ % @ и др. специални символи)."""
    import ctypes
    import ctypes.wintypes
    CF_UNICODETEXT, GMEM_MOVEABLE = 13, 0x0002
    k32, u32 = ctypes.windll.kernel32, ctypes.windll.user32
    k32.GlobalAlloc.restype  = ctypes.c_void_p
    k32.GlobalAlloc.argtypes = [ctypes.wintypes.UINT, ctypes.c_size_t]
    k32.GlobalLock.restype   = ctypes.c_void_p
    k32.GlobalLock.argtypes  = [ctypes.c_void_p]
    k32.GlobalUnlock.argtypes = [ctypes.c_void_p]
    u32.SetClipboardData.argtypes = [ctypes.wintypes.UINT, ctypes.c_void_p]
    data = (text + "\x00").encode("utf-16-le")
    h = k32.GlobalAlloc(GMEM_MOVEABLE, len(data))
    if not h:
        raise OSError("GlobalAlloc failed")
    p = k32.GlobalLock(h)
    ctypes.memmove(p, data, len(data))
    k32.GlobalUnlock(h)
    u32.OpenClipboard(None)
    u32.EmptyClipboard()
    u32.SetClipboardData(CF_UNICODETEXT, h)
    u32.CloseClipboard()


def _dispatch_events(driver, field):
    """Тригерира input/change/blur, за да усети Vue/Quasar промяната."""
    driver.execute_script("""
        var el = arguments[0];
        ['input','change','blur'].forEach(function(e){
            el.dispatchEvent(new Event(e,{bubbles:true,cancelable:true}));});
    """, field)


def _vue_type(driver, field, text: str, clear=True):
    """
    Въвежда текст в Quasar/Vue input.
    Първо опитва РЕАЛНО писане (send_keys) — най-надеждно и задейства v-model.
    Ако стойността не съвпадне (заради специални символи/клавиатурна подредба),
    минава на clipboard paste + native setter като резервен вариант.
    """
    from selenium.webdriver.common.action_chains import ActionChains
    try:
        field.click()
    except Exception:
        driver.execute_script("arguments[0].focus();", field)
    time.sleep(SHORT)
    if clear:
        ActionChains(driver).key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
        time.sleep(0.1)
        ActionChains(driver).send_keys(Keys.DELETE).perform()
        time.sleep(0.1)

    # 1) реално писане
    try:
        field.send_keys(text)
        time.sleep(SHORT)
    except Exception:
        pass
    _dispatch_events(driver, field)
    time.sleep(0.2)

    # 2) ако не е въведено правилно → clipboard + native setter
    if (field.get_attribute("value") or "").strip() != text.strip():
        try:
            ActionChains(driver).key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
            ActionChains(driver).send_keys(Keys.DELETE).perform()
            _copy_to_clipboard(text)
            ActionChains(driver).key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()
            time.sleep(SHORT)
        except Exception:
            pass
        driver.execute_script("""
            var el = arguments[0], val = arguments[1];
            var setter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value').set;
            setter.call(el, val);
        """, field, text)
        _dispatch_events(driver, field)
        time.sleep(SHORT)


def field_input_by_label(driver, label: str, timeout=10):
    """Намира input/textarea на Quasar поле по текста на неговия етикет."""
    xpaths = [
        f"//label[contains(@class,'q-field') and "
        f".//*[contains(@class,'q-field__label') and normalize-space(.)='{label}']]"
        f"//*[self::input or self::textarea]",
        # по-широк вариант — етикетът е някъде в полето
        f"//label[contains(@class,'q-field') and .//*[normalize-space(.)='{label}']]"
        f"//*[self::input or self::textarea]",
    ]
    for xp in xpaths:
        try:
            return WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, xp)))
        except Exception:
            continue
    raise RuntimeError(f"Не намерих поле с етикет '{label}'.")


def detect_date_format(sample: str) -> str:
    """Връща strftime формат според примерна стойност в полето."""
    if not sample:
        return "%Y/%m/%d"
    s = sample.strip()
    if re.match(r"^\d{4}/\d{2}/\d{2}$", s):
        return "%Y/%m/%d"
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return "%Y-%m-%d"
    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", s):
        return "%d.%m.%Y"
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        return "%d/%m/%Y"
    return "%Y/%m/%d"


def to_date(d: str) -> datetime:
    """Парсва конфигурационна дата (ДД.ММ.ГГГГ или ISO) към datetime."""
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(d.strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"Непознат формат за дата: {d}")


def fill_single_date(driver, label: str, date_str: str):
    """Попълва единично дата-поле (напр. „От дата"/„До дата" в справката)."""
    inp = field_input_by_label(driver, label)
    current = inp.get_attribute("value") or ""
    fmt = detect_date_format(current)
    formatted = to_date(date_str).strftime(fmt)
    _vue_type(driver, inp, formatted)
    # затваря евентуален календар
    try:
        inp.send_keys(Keys.ESCAPE)
    except Exception:
        pass
    time.sleep(SHORT)
    print(f"  → Дата '{label}' = {formatted}")


# Селектор за изскачащ попъп (Quasar 2.x): меню / дата-пикер / диалог
POPUP_XP = ("//div[contains(@class,'q-menu')] | //div[contains(@class,'q-date')] "
            "| //div[contains(@class,'q-popup-edit')] | //*[@role='dialog']")


def _field_container(driver, label: str, timeout=10):
    """Връща контейнера (label.q-field) на поле по текста на етикета."""
    xp = (f"//label[contains(@class,'q-field') and "
          f".//*[contains(@class,'q-field__label') and normalize-space(.)='{label}']]")
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.XPATH, xp)))


def _open_date_popup(driver, label: str):
    """Отваря календара на дата-поле (с повторни опити и почистване на стар попъп)."""
    from selenium.webdriver.common.action_chains import ActionChains
    for attempt in range(4):
        # затваря евентуален отворен/остатъчен попъп
        try:
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        except Exception:
            pass
        time.sleep(SHORT)
        try:
            field = _field_container(driver, label, timeout=12)
        except Exception:
            time.sleep(1)
            continue
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", field)
        time.sleep(SHORT)
        for xp in [".//i[contains(@class,'q-icon')]",
                   ".//*[contains(@class,'q-field__append')]",
                   ".//*[contains(@class,'q-field__control')]",
                   ".//input", "."]:
            try:
                el = field.find_element(By.XPATH, xp)
                driver.execute_script("arguments[0].click();", el)
                time.sleep(0.8)
                if driver.find_elements(By.XPATH, POPUP_XP):
                    return
            except Exception:
                continue
        time.sleep(0.8)
    debug_dump(driver, f"date_popup_{label}")
    raise RuntimeError(f"Не успях да отворя календара на поле '{label}'.")


_BG_MONTHS = {"януари": 1, "февруари": 2, "март": 3, "април": 4, "май": 5,
              "юни": 6, "юли": 7, "август": 8, "септември": 9, "октомври": 10,
              "ноември": 11, "декември": 12}


def _qdate_pick(driver, target: datetime):
    """
    Избира дата в Quasar QDate календар (в отворен попъп) чрез стрелките за месец
    (мрежата с дни остава видима) и после клик на деня.
    """
    pop = POPUP_XP
    nav_btn = (f"({pop})//div[contains(@class,'q-date__navigation')]"
               f"//button[not(contains(@class,'q-date__arrow'))]")
    arrow = f"({pop})//button[contains(@class,'q-date__arrow')]"
    target_idx = target.year * 12 + target.month

    for _ in range(60):
        btns = driver.find_elements(By.XPATH, nav_btn)
        if len(btns) < 2:
            break
        mon_txt = (btns[0].text or "").strip().lower()
        yr_txt = re.sub(r"\D", "", btns[1].text or "")
        cur_m = next((v for k, v in _BG_MONTHS.items() if k in mon_txt), None)
        if not cur_m or not yr_txt:
            break
        cur_idx = int(yr_txt) * 12 + cur_m
        if cur_idx == target_idx:
            break
        arrows = driver.find_elements(By.XPATH, arrow)
        if len(arrows) < 2:
            break
        # arrows[0]=месец назад, arrows[1]=месец напред (годината се сменя сама)
        driver.execute_script("arguments[0].click();",
                              arrows[1] if target_idx > cur_idx else arrows[0])
        time.sleep(0.35)

    # клик на деня (само дни от текущия месец, не „out" дните)
    day_xp = (f"({pop})//div[contains(@class,'q-date__calendar-item') and "
              f"not(contains(@class,'q-date__calendar-item--out'))]"
              f"//*[self::button or self::div][normalize-space(.)='{target.day}']")
    el = WebDriverWait(driver, 6).until(EC.element_to_be_clickable((By.XPATH, day_xp)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    driver.execute_script("arguments[0].click();", el)
    time.sleep(0.4)


def fill_date_range(driver, label: str, date_from: str, date_to: str):
    """
    Попълва период в „Разширено търсене" (поле като „Създадена"):
    отваря календара → радио „Период" → клик на начална и крайна дата → ИЗБЕРИ.
    """
    _open_date_popup(driver, label)

    # избира радио „Период"
    try:
        el = driver.find_element(
            By.XPATH, f"({POPUP_XP})//*[contains(@class,'q-radio') and "
                      f".//*[normalize-space(.)='Период']]")
        driver.execute_script("arguments[0].click();", el)
        time.sleep(SHORT)
    except Exception:
        pass

    # кликва начална, после крайна дата в календара
    _qdate_pick(driver, to_date(date_from))
    _qdate_pick(driver, to_date(date_to))

    # бутон ИЗБЕРИ
    for txt in ["ИЗБЕРИ", "Избери", "OK", "ОК"]:
        try:
            btn = driver.find_element(
                By.XPATH, f"({POPUP_XP})//button[.//*[normalize-space(.)='{txt}'] "
                          f"or normalize-space(.)='{txt}']")
            driver.execute_script("arguments[0].click();", btn)
            break
        except Exception:
            continue
    time.sleep(MED)
    print(f"  → Период '{label}': {date_from} – {date_to}")


def select_quasar_dropdown(driver, label: str, option: str):
    """Избира стойност от Quasar q-select по етикет на полето и текст на опцията."""
    inp = field_input_by_label(driver, label)
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
    try:
        inp.click()
    except Exception:
        driver.execute_script("arguments[0].click();", inp)
    time.sleep(MED)
    opt_xp = ("//div[contains(@class,'q-menu')]//*[contains(@class,'q-item') and "
              f"normalize-space(.)='{option}']")
    try:
        el = WebDriverWait(driver, 6).until(
            EC.element_to_be_clickable((By.XPATH, opt_xp)))
        driver.execute_script("arguments[0].click();", el)
    except Exception:
        debug_dump(driver, f"dropdown_{label}")
        raise RuntimeError(f"Не намерих опция '{option}' в падащото меню '{label}'.")
    time.sleep(SHORT)
    print(f"  → '{label}' = {option}")


def select_company(driver, label: str, company: str) -> bool:
    """Избира компания от автодовършващото поле: кликва → пише → избира от списъка."""
    inp = field_input_by_label(driver, label)
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
    _vue_type(driver, inp, company)
    time.sleep(MED)
    up = ("translate(normalize-space(.),"
          "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ')")
    opt_xp = (f"//div[contains(@class,'q-menu')]//*[contains(@class,'q-item')]"
              f"[contains({up},'{company.upper()}')]")
    try:
        el = WebDriverWait(driver, 6).until(
            EC.element_to_be_clickable((By.XPATH, opt_xp)))
        driver.execute_script("arguments[0].click();", el)
        print(f"  → Компания избрана: {company}")
        time.sleep(SHORT)
        return True
    except Exception:
        try:
            first = driver.find_element(
                By.XPATH, "//div[contains(@class,'q-menu')]//*[contains(@class,'q-item')][1]")
            driver.execute_script("arguments[0].click();", first)
            print(f"  ⚠ '{company}' — избран първият резултат (провери ръчно).")
            return True
        except Exception:
            debug_dump(driver, f"company_{company}")
            print(f"  ⚠ Не успях да избера компания '{company}'.")
            return False


def set_status_checkboxes(driver, statuses, clear_first=True):
    """Маркира чекбоксове за статус по техния текст (в групата „Статус")."""
    all_status = ["Очаквана", "Настанена", "Изпратена", "Анулирана", "Непристигнала"]
    if clear_first:
        for st in all_status:
            _set_checkbox(driver, st, want=False)
    for st in statuses:
        if not _set_checkbox(driver, st, want=True):
            print(f"  ⚠ Не намерих чекбокс за статус: {st}")


def _set_checkbox(driver, text: str, want: bool) -> bool:
    """Поставя чекбокс (Quasar q-checkbox) по етикет в желано състояние."""
    xp = (f"//*[contains(@class,'q-checkbox')][.//*[normalize-space(.)='{text}']]")
    try:
        cb = driver.find_element(By.XPATH, xp)
    except Exception:
        try:
            cb = driver.find_element(
                By.XPATH, f"//*[normalize-space(.)='{text}']"
                          f"/preceding::*[contains(@class,'q-checkbox')][1]")
        except Exception:
            return False
    checked = (cb.get_attribute("aria-checked") == "true")
    if checked != want:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb)
        driver.execute_script("arguments[0].click();", cb)
        time.sleep(0.15)
    return True


def click_search(driver):
    """
    Кликва бутона за търсене НА ФОРМАТА (червената плаваща лупа долу/горе вдясно),
    а НЕ глобалното „Бързо търсене" в горната лента (то е в q-header).
    """
    selectors = [
        # 1) плаващ бутон (q-page-sticky) с лупа — това е бутонът на формата
        "//*[contains(@class,'q-page-sticky')]//button[.//i[normalize-space(.)='search']]",
        # 2) кръгъл/fab бутон с лупа, извън горната лента
        "//button[(contains(@class,'q-btn--round') or contains(@class,'q-btn--fab')) "
        "and .//i[normalize-space(.)='search'] "
        "and not(ancestor::*[contains(@class,'q-header')])]",
        # 3) всеки бутон с лупа, който НЕ е в горната лента/header
        "//button[.//i[normalize-space(.)='search'] "
        "and not(ancestor::*[contains(@class,'q-header')]) and not(ancestor::header)]",
    ]
    for sel in selectors:
        try:
            btn = WebDriverWait(driver, 6).until(EC.element_to_be_clickable((By.XPATH, sel)))
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(MED)
            return
        except Exception:
            continue
    debug_dump(driver, "search_button")
    raise RuntimeError("Не намерих бутона за търсене на формата (лупата).")


def wait_for_download(directory: Path, timeout: int = 90) -> Path:
    """Изчаква последния изтеглен .xls/.xlsx файл (в нашата + системната Downloads)."""
    dirs = [directory, Path.home() / "Downloads"]
    start = time.time()
    end = start + timeout
    while time.time() < end:
        cand = []
        for d in dirs:
            try:
                for f in d.iterdir():
                    if (f.suffix.lower() in (".xlsx", ".xls")
                            and not f.name.endswith(".crdownload")
                            and f.stat().st_mtime >= start - 5):
                        cand.append(f)
            except Exception:
                continue
        cand.sort(key=lambda f: f.stat().st_mtime, reverse=True)
        if cand:
            latest = cand[0]
            size = latest.stat().st_size
            time.sleep(2)
            if latest.stat().st_size == size and size > 0:
                return latest
        time.sleep(1)
    raise TimeoutError(f"Файлът не беше изтеглен за {timeout} секунди.")


def export_excel(driver, dest_name: str) -> Path:
    """
    Кликва бутона за сваляне (Font Awesome икона) → меню → „Excel".
    Опитва няколко кандидат-бутона и потвърждава, че се е отворило меню с „Excel".
    """
    # Бутонът за сваляне е Quasar QFab (икона „download"), който при клик
    # разгъва под-действия CSV / Excel (q-fab__label). Отваряме QFab, после
    # кликаме действието „Excel".
    triggers = [
        "//*[contains(@class,'q-fab__icon-holder')]",
        "//button[.//i[normalize-space(.)='download' or normalize-space(.)='file_download' "
        "or normalize-space(.)='get_app' or normalize-space(.)='save_alt' "
        "or normalize-space(.)='cloud_download']]",
        "//*[contains(@class,'q-fab')]//button",
        "//button[.//i[contains(@class,'fa-download') "
        "or contains(@class,'fa-cloud-download') or contains(@class,'fa-file-export')]]",
    ]
    excel_lbl = "//*[contains(@class,'q-fab__label') and normalize-space(.)='Excel']"

    clicked = False       # кликнахме опцията „Excel" в меню
    triggered = False     # кликнахме поне един бутон за сваляне
    for txp in triggers:
        for trg in driver.find_elements(By.XPATH, txp):
            try:
                if not trg.is_displayed():
                    continue
                driver.execute_script("arguments[0].click();", trg)
                triggered = True
                time.sleep(MED)
            except Exception:
                continue
            labels = [e for e in driver.find_elements(By.XPATH, excel_lbl)
                      if e.is_displayed()]
            if not labels:
                labels = [e for e in driver.find_elements(
                    By.XPATH, "//*[normalize-space(.)='Excel']") if e.is_displayed()]
            if labels:
                target = labels[0]
                try:
                    target = target.find_element(By.XPATH, "ancestor-or-self::button[1]")
                except Exception:
                    pass
                driver.execute_script("arguments[0].click();", target)
                clicked = True
                break
        if clicked:
            break
    if not (clicked or triggered):
        debug_dump(driver, "download_button")
        raise RuntimeError("Не намерих бутон за сваляне.")

    # Изчаква файла — важи и при меню „Excel", и при директно сваляне.
    try:
        downloaded = wait_for_download(DOWNLOAD_DIR)
    except TimeoutError:
        debug_dump(driver, "download_button")
        raise RuntimeError("Натиснах сваляне, но файл не се появи (Excel опция/директно сваляне).")
    dest = WORK_DIR / dest_name
    if dest.exists():
        dest.unlink()
    shutil.move(str(downloaded), str(dest))
    print(f"  → Запазено: {dest.name}")
    return dest


# ═════════════════════════════════════════════════════════════
#  ЛОГИН
# ═════════════════════════════════════════════════════════════

def _logged_in(driver) -> bool:
    """
    Логнати сме, ако НЯКОЙ отворен таб е на вътрешен екран (/hotel/<номер>).
    Обхожда всички табове и се ПРЕВКЛЮЧВА на правилния (важно при закачен режим,
    където може да има няколко таба).
    """
    try:
        handles = driver.window_handles
    except Exception:
        return False
    for h in handles:
        try:
            driver.switch_to.window(h)
            if re.search(r"/hotel/\d+", driver.current_url or ""):
                return True
        except Exception:
            continue
    return False


def _click_next(driver):
    """Кликва NEXT (или Enter)."""
    for sel in [
        "//button[normalize-space(.)='NEXT' or normalize-space(.)='Next']",
        "//button[.//*[normalize-space(.)='NEXT' or normalize-space(.)='Next']]",
    ]:
        try:
            btn = WebDriverWait(driver, 6).until(EC.element_to_be_clickable((By.XPATH, sel)))
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(MED)
            return
        except Exception:
            continue
    try:
        driver.switch_to.active_element.send_keys(Keys.RETURN)
    except Exception:
        pass
    time.sleep(MED)


def login(driver):
    """
    Влиза в ClockPMS+. При запазен профил обикновено има само екран за ПАРОЛА
    (серийният ключ и потребителят са запомнени). Кодът е адаптивен:
      • ако вече сме логнати → излиза
      • ако има поле за парола → попълва паролата → NEXT
      • ако започва от сериен ключ → серия → потребител → парола
    """
    def _wait_until(cond, timeout=20):
        end = time.time() + timeout
        while time.time() < end:
            try:
                if cond():
                    return True
            except Exception:
                pass
            time.sleep(0.5)
        return False

    print("[LOGIN] Проверявам входа в ClockPMS+...")
    time.sleep(1)

    # ── РЕЖИМ „закачане": потребителят се логва РЪЧНО ──
    if ATTACH_TO_RUNNING:
        if _logged_in(driver):
            print("[LOGIN] ✓ Вече си логнат.")
            return
        print("\n" + "!" * 60)
        print("Влез в ClockPMS+ в отворения Chrome (ако още не си).")
        print("Скриптът чака до 5 минути и продължава сам...")
        print("!" * 60)
        if not _wait_until(_logged_in, 300):
            debug_dump(driver, "attach_not_logged_in")
            raise RuntimeError("Не открих логнат ClockPMS+ в отворения Chrome.")
        print("[LOGIN] ✓ Открих логната сесия.")
        return

    # ── РЕЖИМ „скриптът пуска Chrome": автоматичен логин ──
    driver.get(url("home"))
    time.sleep(4)
    if _logged_in(driver):
        print("[LOGIN] ✓ Вече сме логнати.")
        return

    # ── Автоматичен опит за логин (толерантен — при проблем минаваме на ръчен) ──
    try:
        # 1) Сериен ключ (екран „Serial")
        if "Serial" in driver.page_source and "Password" not in driver.page_source:
            print("  [serial] Въвеждам сериен ключ...")
            sf = driver.find_element(By.CSS_SELECTOR, "input.q-field__native")
            _vue_type(driver, sf, SERIAL_KEY)
            if not (sf.get_attribute("value") or "").strip():
                print("    ⚠ Серийното поле изглежда празно след въвеждане.")
            _click_next(driver)
            _wait_until(lambda: "Serial" not in driver.page_source, 15)

        # 2) Потребител (ако се иска преди паролата)
        if "Password" not in driver.page_source:
            text_inputs = [e for e in driver.find_elements(By.CSS_SELECTOR, "input[type='text']")
                           if e.is_displayed()]
            if text_inputs:
                print("  [user] Въвеждам потребител...")
                _vue_type(driver, text_inputs[0], USERNAME)
                _click_next(driver)
                _wait_until(lambda: "Password" in driver.page_source
                            or driver.find_elements(By.CSS_SELECTOR, "input[type='password']"), 15)

        # 3) Парола
        pwd = driver.find_elements(By.CSS_SELECTOR, "input[type='password']")
        if pwd and pwd[0].is_displayed():
            print("  [password] Въвеждам парола...")
            _vue_type(driver, pwd[0], PASSWORD)
            _click_next(driver)
            time.sleep(3)
    except Exception as e:
        print(f"  ⚠ Автоматичният логин не успя ({e}). Минавам на ръчен вход.")
        debug_dump(driver, "login_auto_failed")

    if any(t in driver.page_source for t in
           ("Unsuccessful", "Failed to authenticate", "incorrect", "невалид")):
        debug_dump(driver, "login_failed")
        print("  ⚠ Възможен неуспешен вход — провери USERNAME/PASSWORD.")

    # ── Ако не сме логнати → изчакваме РЪЧЕН вход в отворения прозорец ──
    if not _logged_in(driver):
        print("\n" + "!" * 60)
        print("РЪЧЕН ВХОД: довърши логина в отворения Chrome прозорец.")
        print("(Това е нужно само първия път — после профилът го помни.)")
        print("Скриптът ще изчака до 5 минути...")
        print("!" * 60)
        if not _wait_until(_logged_in, timeout=300):
            debug_dump(driver, "login_unconfirmed")
            raise RuntimeError("Логинът не беше потвърден (изтекоха 5 минути).")

    print("[LOGIN] ✓ Влязохме в ClockPMS+.")


# ═════════════════════════════════════════════════════════════
#  PHASE 1 — WEB SCRAPING
# ═════════════════════════════════════════════════════════════

def step1_1_new_reservations(driver) -> Path:
    """1.1 Нови резервации → New_Reservations.xlsx"""
    print("\n[1.1] Нови резервации...")
    driver.get(url("bookings/search/advanced"))
    time.sleep(MED + 1)
    fill_date_range(driver, DATE_FIELD_LABEL, NEW_RES_FROM, NEW_RES_TO)
    set_status_checkboxes(driver, NEW_RES_STATUSES, clear_first=True)
    click_search(driver)
    WebDriverWait(driver, 20).until(EC.url_contains("/result"))
    time.sleep(MED)
    return export_excel(driver, "New_Reservations.xlsx")


def step1_2_cancellations(driver) -> Path:
    """1.2 Анулации → Cancellations.xlsx"""
    print("\n[1.2] Анулации...")
    driver.get(url("bookings/search/advanced"))
    time.sleep(MED + 1)
    fill_date_range(driver, DATE_FIELD_LABEL, CANCEL_FROM, CANCEL_TO)
    set_status_checkboxes(driver, [CANCEL_STATUS], clear_first=True)
    click_search(driver)
    WebDriverWait(driver, 20).until(EC.url_contains("/result"))
    time.sleep(MED)
    return export_excel(driver, "Cancellations.xlsx")


def _match_company(name: str):
    """Връща ключа от COMPANY_SHEET_MAP при точно или ≥90% съвпадение, иначе None."""
    import difflib
    name = (name or "").strip()
    if not name:
        return None
    if name in COMPANY_SHEET_MAP:
        return name
    best, score = None, 0.0
    for key in COMPANY_SHEET_MAP:
        r = difflib.SequenceMatcher(None, name.upper(), key.upper()).ratio()
        if r > score:
            best, score = key, r
    return best if score >= 0.90 else None


# Фирми/записи, които НЕ са турагенции и се прескачат изцяло
SKIP_COMPANIES = {
    "TOPOLA SKIES", "INDIVIDUAL",
    "ЕКСПРЕС ГАРАНЦИОН - ООД", "АКВАТЕК ООД", "ТОПОЛА СКАЙС МЕНИДЖМЪНТ ООД",
}


def _sanitize_sheet_name(name: str) -> str:
    """Прави валидно име на Excel лист (≤31 символа, без : \\ / ? * [ ])."""
    s = re.sub(r"[:\\/?*\[\]]", " ", (name or "").strip())
    s = re.sub(r"\s+", " ", s).strip()
    return (s[:31] or "Sheet")


def _sheet_for_company(company: str) -> str:
    """Връща листа за фирмата: от речника (точно/≥90%) или ново име по фирмата."""
    key = _match_company(company)
    if key:
        return COMPANY_SHEET_MAP[key]
    return _sanitize_sheet_name(company)


def _is_skipped(company: str) -> bool:
    c = (company or "").strip()
    return c == "" or c.upper() in {s.upper() for s in SKIP_COMPANIES}


def step1_3_missing_companies(new_res_path=None) -> list:
    """
    1.3 Компании за справка = ВСИЧКИ фирми от първата справка от PMS
    (New_Reservations.xlsx, колона „Компания"), без изрично прескочените.
    """
    print("\n[1.3] Определям компании за справка (от New_Reservations.xlsx)...")
    companies, skipped = [], []
    if not (new_res_path and Path(new_res_path).exists()):
        print("  ⚠ Липсва New_Reservations.xlsx — няма компании.")
        return []
    try:
        dfn = pd.read_excel(new_res_path)
    except Exception as e:
        print(f"  ⚠ Грешка при четене на новите резервации: {e}")
        return []
    ccol = next((c for c in dfn.columns
                 if isinstance(c, str) and c.strip().lower() == "компания"), None)
    if not ccol:
        print("  ⚠ Няма колона 'Компания' в новите резервации.")
        return []

    seen = set()
    for v in dfn[ccol].dropna().astype(str):
        for part in v.split(","):        # клетка може да съдържа няколко фирми
            name = part.strip()
            if not name or name in seen:
                continue
            seen.add(name)
            if _is_skipped(name):
                skipped.append(name)
            else:
                companies.append(name)

    if skipped:
        print(f"  • Прескочени (не са турагенции): {skipped}")
    new_sheets = [c for c in companies if not _match_company(c)]
    if new_sheets:
        print(f"  • Ще се създадат НОВИ листове за: {new_sheets}")
    print(f"  → {len(companies)} компании за справка: {companies}")
    return companies


def step1_4_company_reports(driver, companies: list) -> dict:
    """1.4 Справка „Начисления от резервации на компания по референция" по компания."""
    print("\n[1.4] Справки по компания...")
    results = {}
    for company in companies:
        if _is_skipped(company):
            continue
        sheet = _sheet_for_company(company)
        print(f"  → {company}  (лист: {sheet})")
        # нулира формата: смяна на route (Начало → справката) пресъздава чиста форма
        driver.get(url("home"))
        time.sleep(MED)
        driver.get(url("reports/company-booking-charge-by-reference"))
        time.sleep(MED + 1)
        # ако формата е в „резултатен" режим (няма поле Компания) — отваря я с молива
        try:
            field_input_by_label(driver, "Компания", timeout=4)
        except Exception:
            _open_report_form(driver)
        try:
            select_quasar_dropdown(driver, "Групирай по", REPORT_GROUPBY)
        except Exception:
            pass
        fill_single_date(driver, "От дата", COMPANY_FROM)
        fill_single_date(driver, "До дата", COMPANY_TO)
        try:
            select_quasar_dropdown(driver, "Стойност", REPORT_VALUE)
        except Exception:
            pass
        if not select_company(driver, "Компания", company):
            continue
        click_search(driver)
        _wait_report_ready(driver)
        try:
            dest = export_excel(driver, f"Company_{_sanitize_sheet_name(sheet)}.xlsx")
            results[company] = dest
        except Exception as e:
            print(f"  ⚠ Неуспешен експорт за '{company}': {e}")
            continue
    return results


def _open_report_form(driver):
    """Отваря формата с критерии (бутона-молив „edit"), ако сме в резултатен режим."""
    xps = [
        "//button[.//i[normalize-space(.)='edit' or normalize-space(.)='create' "
        "or normalize-space(.)='mode_edit']]",
        "//button[.//i[contains(@class,'fa-pencil') or contains(@class,'fa-edit')]]",
    ]
    for xp in xps:
        for b in driver.find_elements(By.XPATH, xp):
            try:
                if b.is_displayed():
                    driver.execute_script("arguments[0].click();", b)
                    time.sleep(MED)
                    return True
            except Exception:
                continue
    return False


def _wait_report_ready(driver, timeout=180):
    """Изчаква справката да се генерира (изчезване на „Данните се обработват…")."""
    end = time.time() + timeout
    time.sleep(MED)
    while time.time() < end:
        try:
            src = driver.page_source
        except Exception:
            src = ""
        if "обработват" not in src and "Обработват" not in src:
            time.sleep(MED)
            return True
        time.sleep(2)
    print("  ⚠ Справката още се обработва след изчакване — продължавам.")
    return False


# ═════════════════════════════════════════════════════════════
#  PHASE 2 — EXCEL PROCESSING
# ═════════════════════════════════════════════════════════════

def _shift_formula_row(formula: str, old_row: int, new_row: int) -> str:
    """Адаптира относителните референции old_row → new_row (без да пипа $абсолютни)."""
    def repl(m):
        prefix, col, dollar2, row = m.group(1), m.group(2), m.group(3), int(m.group(4))
        if dollar2 or row != old_row:   # $A$1 → не пипаме реда
            return m.group(0)
        return f"{prefix}{col}{dollar2}{new_row}"
    return re.sub(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", repl, formula)


def step2_1_integrate_new_reservations(wb, new_res_path: Path) -> int:
    print("\n[2.1] Добавяне на нови резервации...")
    ws = wb[SHEET_RES]
    df_new = pd.read_excel(new_res_path)
    if "Номер" in df_new.columns:
        df_new["Номер"] = df_new["Номер"].astype(str).str.replace("#", "", regex=False)

    last_row = ws.max_row
    while last_row > 1 and all(ws.cell(last_row, c).value is None
                               for c in range(1, ws.max_column + 1)):
        last_row -= 1

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    skip_cols = {"Начисления за стая", "ADR"}
    added = 0
    for _, row_data in df_new.iterrows():
        prev_row = last_row
        last_row += 1
        for c, header in enumerate(headers, start=1):
            if header in skip_cols:
                continue
            if header and header in df_new.columns:
                ws.cell(last_row, c, row_data.get(header))
        for c, header in enumerate(headers, start=1):
            if header in skip_cols:
                continue
            pv = ws.cell(prev_row, c).value
            if isinstance(pv, str) and pv.startswith("="):
                ws.cell(last_row, c, _shift_formula_row(pv, prev_row, last_row))
        added += 1
    print(f"  → Добавени {added} реда.")
    return added


def _find_number_column(df: pd.DataFrame):
    """Намира колоната с номер на резервация в експорт от PMS (#, Номер, №…)."""
    for c in df.columns:
        if isinstance(c, str) and (c.strip() in ("#", "№", "Номер")
                                   or "омер на резерв" in c.lower()):
            return c
    return None


def step2_2_integrate_cancellations(wb, cancel_path: Path):
    print("\n[2.2] Интеграция на анулации...")
    ws = wb[SHEET_RES]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    num_col = headers.get("Номер")
    anul_col = headers.get("Анулирана") or column_index_from_string("AL")
    if not num_col:
        print("  ⚠ Няма колона 'Номер' в листа — пропускам.")
        return
    df_cancel = pd.read_excel(cancel_path)
    ncol = _find_number_column(df_cancel)
    if not ncol:
        print(f"  ⚠ Няма колона с номер в анулациите ({list(df_cancel.columns)[:6]}…).")
        return
    cancelled = set(df_cancel[ncol].dropna().astype(str)
                    .str.replace("#", "", regex=False).str.strip())
    n = 0
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, num_col).value
        if v is not None and str(v).strip() in cancelled:
            ws.cell(row, anul_col, "Анулирана")
            n += 1
    print(f"  → Маркирани {n} анулирани реда (колона '{get_column_letter(anul_col)}').")


def step2_3_update_blue_sheets(wb, company_files: dict) -> list:
    print("\n[2.3] Обновяване на сини листове...")
    updated = []
    for company, path in company_files.items():
        sheet = _sheet_for_company(company)
        if sheet not in wb.sheetnames:
            wb.create_sheet(sheet)          # нов лист за непозната фирма
            print(f"  + Създаден нов лист '{sheet}' за {company}")
        print(f"  → {company} → лист '{sheet}'")
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        # Копира справката ВЕРБАТИМ (за да се запазят колони C=номер и L=Стойност)
        src = load_workbook(path, data_only=True)
        sws = src.active
        for r, rowvals in enumerate(sws.iter_rows(values_only=True), start=1):
            for c, val in enumerate(rowvals, start=1):
                ws.cell(r, c, val)
        src.close()
        # Колона L (Стойност): „502,29 EUR" → число в лева
        l = column_index_from_string("L")
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(r, l)
            if cell.value is None:
                continue
            raw = (str(cell.value).strip().replace(" EUR", "").replace(" eur", "")
                   .replace(" ", "").replace(" ", "").replace(",", "."))
            try:
                cell.value = round(float(raw) * EUR_TO_BGN, 2)
            except ValueError:
                pass
        updated.append(sheet)
    print(f"  → Обновени листове: {updated}")
    return updated


def step2_4_inject_lookup_formulas(wb):
    print("\n[2.4] Инжектиране на XLOOKUP формули...")
    ws = wb[SHEET_RES]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    num_col = headers.get("Номер", 1)
    comp_col = headers.get("Компания")
    charge_col = headers.get("Начисления за стая")
    adr_col = headers.get("ADR")
    if not (comp_col and charge_col):
        print("  ⚠ Липсват колони 'Компания'/'Начисления за стая' — пропускам.")
        return
    num_l = get_column_letter(num_col)
    charge_l = get_column_letter(charge_col)
    nights_col = headers.get("Престой") or headers.get("Нощувки")
    injected = 0
    for row in range(2, ws.max_row + 1):
        comp = ws.cell(row, comp_col).value
        if comp is None or _is_skipped(str(comp)):
            continue
        sheet = _sheet_for_company(str(comp).strip())
        if not sheet:
            continue
        if ws.cell(row, charge_col).value in (None, ""):
            # В справката: номер на резервация = колона C, Стойност = колона L
            ws.cell(row, charge_col).value = (
                f"=IFERROR(XLOOKUP({num_l}{row},'{sheet}'!$C:$C,'{sheet}'!$L:$L),\"\")")
            injected += 1
        if adr_col and ws.cell(row, adr_col).value in (None, ""):
            if nights_col:
                nl = get_column_letter(nights_col)
                ws.cell(row, adr_col).value = f"=IFERROR({charge_l}{row}/{nl}{row},\"\")"
            else:
                ws.cell(row, adr_col).value = f"={charge_l}{row}"
    print(f"  → Инжектирани {injected} XLOOKUP формули.")


# ═════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════

def main():
    start = datetime.now()
    print("=" * 60)
    print("ClockPMS+ RPA + Excel Processing")
    print(f"Стартирано: {start:%Y-%m-%d %H:%M:%S}")
    print("=" * 60)

    driver = attach_driver() if ATTACH_TO_RUNNING else build_driver()
    new_res_path = cancel_path = None
    company_files = {}
    try:
        login(driver)
        new_res_path  = step1_1_new_reservations(driver)
        cancel_path   = step1_2_cancellations(driver)
        companies     = step1_3_missing_companies(new_res_path)
        company_files = step1_4_company_reports(driver, companies)
    except Exception as e:
        print(f"\n❌ ГРЕШКА във фаза 1: {e}")
        traceback.print_exc()
        try:
            debug_dump(driver, "phase1_error")
        except Exception:
            pass
        print("Виж папка debug/ за снимка и HTML на страницата.")
    finally:
        if ATTACH_TO_RUNNING:
            print("\n[BROWSER] Оставям Chrome отворен (закачен режим).")
        else:
            driver.quit()
            print("\n[BROWSER] Затворен.")

    if not (new_res_path or cancel_path or company_files):
        print("Няма свалени данни — спирам преди Excel обработката.")
        return

    print("\n[PHASE 2] Обработка на Excel...")
    try:
        print(f"  Копирам {MASTER_FILE.name} → {OUTPUT_FILE.name} ...")
        shutil.copy(str(MASTER_FILE), str(OUTPUT_FILE))
        print("  Зареждам файла (може да отнеме време при голям файл)...")
        wb = load_workbook(OUTPUT_FILE, keep_links=False)
        print("  Файлът е зареден.")
        added = 0
        if new_res_path:
            added = step2_1_integrate_new_reservations(wb, new_res_path)
        if cancel_path:
            step2_2_integrate_cancellations(wb, cancel_path)
        updated = step2_3_update_blue_sheets(wb, company_files) if company_files else []
        step2_4_inject_lookup_formulas(wb)
        print("  Записвам резултата...")
        wb.save(OUTPUT_FILE)
        print(f"\n[SAVE] Записано: {OUTPUT_FILE.name}")
    except PermissionError:
        print("\n❌ Нямам достъп до файла за запис. Затвори "
              f"'{OUTPUT_FILE.name}' и '{MASTER_FILE.name}' в Excel и пусни пак.")
        return
    except Exception as e:
        print(f"\n❌ ГРЕШКА във Фаза 2: {e}")
        traceback.print_exc()
        return

    print("\n" + "=" * 60)
    print("РЕЗУЛТАТ:")
    print(f"  Нови реда в '{SHEET_RES}': {added}")
    print(f"  Обновени сини листове ({len(updated)}): {', '.join(updated)}")
    print(f"  Изходен файл: {OUTPUT_FILE}")
    elapsed = (datetime.now() - start).seconds
    print(f"  Общо време: {elapsed // 60}м {elapsed % 60}с")
    print("=" * 60)


if __name__ == "__main__":
    main()
